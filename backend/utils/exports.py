import io
import pandas as pd
from datetime import datetime
from fastapi.responses import StreamingResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
import os
import logging

logger = logging.getLogger(__name__)

# ─── Arabic Font & RTL Setup ────────────────────────────────────────────────

ARABIC_FONT_PATH = "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf"
ARABIC_FONT_BOLD_PATH = "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf"
ARABIC_FONT_NAME = "NotoNaskhArabic"
ARABIC_FONT_BOLD_NAME = "NotoNaskhArabic-Bold"
_arabic_font_registered = False


def _register_arabic_font():
    """Register Noto Naskh Arabic font once with ReportLab."""
    global _arabic_font_registered
    if _arabic_font_registered:
        return True
    try:
        if os.path.exists(ARABIC_FONT_PATH):
            pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, ARABIC_FONT_PATH))
            if os.path.exists(ARABIC_FONT_BOLD_PATH):
                pdfmetrics.registerFont(TTFont(ARABIC_FONT_BOLD_NAME, ARABIC_FONT_BOLD_PATH))
            _arabic_font_registered = True
            return True
    except Exception:
        pass
    return False


def _reshape_arabic(text: str) -> str:
    """
    Reshape and apply BiDi algorithm to Arabic text for correct rendering in ReportLab.
    ReportLab renders text left-to-right internally, so Arabic must be pre-processed.
    """
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)


def _has_arabic(text: str) -> bool:
    """Quick check if a string contains Arabic characters."""
    try:
        for ch in str(text):
            if '\u0600' <= ch <= '\u06FF' or '\u0750' <= ch <= '\u077F':
                return True
    except Exception:
        pass
    return False


def _process_cell(cell_value, arabic_available: bool) -> str:
    """Process a cell value: reshape Arabic text if font is available."""
    text = str(cell_value) if cell_value is not None else ''
    if arabic_available and _has_arabic(text):
        return _reshape_arabic(text)
    return text


def generate_excel(data: list, columns: list, sheet_name: str = "Report") -> io.BytesIO:
    """
    Generate Excel file from list of dictionaries
    data: List of dicts
    columns: List of column names (headers)
    """
    df = pd.DataFrame(data)
    
    # Filter/Order columns if specific columns requested
    if columns:
        # Only keep columns that exist in data
        valid_cols = [c for c in columns if c in df.columns]
        if valid_cols:
            df = df[valid_cols]
            
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Auto-adjust column width
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(col))
            )) + 1
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)
            
    output.seek(0)
    return output

def generate_pdf(data: list, title: str, columns: list = None, rtl: bool = None,
                  orientation: str = "auto", company_name: str = None,
                  subtitle: str = None, chart_image: bytes = None) -> io.BytesIO:
    """
    Generate a PDF report with full Arabic/RTL support, page numbers,
    company header, and optional chart embedding.
    
    data: List of dicts or List of Lists (first element = header if columns=None)
    title: Report Title (can contain Arabic)
    columns: Headers (if data is list of dicts)
    rtl: Force RTL mode. If None, auto-detect from title/data.
    orientation: "portrait", "landscape", or "auto" (auto picks landscape for >6 cols)
    company_name: Company name for header
    subtitle: Optional subtitle line (e.g. date range)
    chart_image: Optional chart image bytes (PNG) to embed before table
    """
    arabic_available = _register_arabic_font()
    
    # Auto-detect RTL
    if rtl is None:
        rtl = arabic_available and (_has_arabic(title) or (
            data and isinstance(data[0], (list, dict)) and
            any(_has_arabic(str(v)) for row in data[:3]
                for v in (row.values() if isinstance(row, dict) else row))
        ))
    
    # Auto-detect orientation
    col_count_hint = 0
    if columns:
        col_count_hint = len(columns)
    elif data and isinstance(data[0], dict):
        col_count_hint = len(data[0])
    elif data and isinstance(data[0], list):
        col_count_hint = len(data[0])
    
    if orientation == "auto":
        orientation = "landscape" if col_count_hint > 6 else "portrait"
    
    page_size = landscape(A4) if orientation == "landscape" else A4
    
    buffer = io.BytesIO()
    
    # Page number callback
    def _add_page_number(canvas, doc):
        canvas.saveState()
        page_num = canvas.getPageNumber()
        if rtl and arabic_available:
            canvas.setFont(ARABIC_FONT_NAME, 8)
            footer_text = _reshape_arabic(f"صفحة {page_num}")
        else:
            canvas.setFont('Helvetica', 8)
            footer_text = f"Page {page_num}"
        canvas.setFillColor(colors.HexColor('#9CA3AF'))
        canvas.drawCentredString(page_size[0] / 2, 20, footer_text)
        # Date stamp
        date_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        if rtl:
            canvas.drawString(40, 20, date_str)
        else:
            canvas.drawRightString(page_size[0] - 40, 20, date_str)
        canvas.restoreState()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )
    elements = []
    styles = getSampleStyleSheet()
    
    if rtl and arabic_available:
        title_style = ParagraphStyle(
            'ArabicTitle',
            fontName=ARABIC_FONT_BOLD_NAME if os.path.exists(ARABIC_FONT_BOLD_PATH) else ARABIC_FONT_NAME,
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=4,
            leading=28,
            textColor=colors.HexColor('#1E3A5F'),
        )
        company_style = ParagraphStyle(
            'ArabicCompany',
            fontName=ARABIC_FONT_BOLD_NAME if os.path.exists(ARABIC_FONT_BOLD_PATH) else ARABIC_FONT_NAME,
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=4,
            leading=22,
            textColor=colors.HexColor('#374151'),
        )
        subtitle_style = ParagraphStyle(
            'ArabicSubtitle',
            fontName=ARABIC_FONT_NAME,
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=8,
            leading=16,
            textColor=colors.HexColor('#6B7280'),
        )
        normal_style = ParagraphStyle(
            'ArabicNormal',
            fontName=ARABIC_FONT_NAME,
            fontSize=10,
            alignment=TA_RIGHT,
            spaceAfter=6,
            leading=16,
        )
        header_font = ARABIC_FONT_BOLD_NAME if os.path.exists(ARABIC_FONT_BOLD_PATH) else ARABIC_FONT_NAME
        body_font = ARABIC_FONT_NAME
    else:
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=4,
            textColor=colors.HexColor('#1E3A5F'),
        )
        company_style = ParagraphStyle(
            'CompanyName',
            parent=styles['Heading2'],
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=4,
            textColor=colors.HexColor('#374151'),
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=8,
            textColor=colors.HexColor('#6B7280'),
        )
        normal_style = styles['Normal']
        header_font = 'Helvetica-Bold'
        body_font = 'Helvetica'
    
    # Company header
    if company_name:
        elements.append(Paragraph(_process_cell(company_name, rtl and arabic_available), company_style))
    
    # Title
    elements.append(Paragraph(_process_cell(title, rtl and arabic_available), title_style))
    
    # Subtitle / date range
    if subtitle:
        elements.append(Paragraph(_process_cell(subtitle, rtl and arabic_available), subtitle_style))
    else:
        date_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        gen_label = 'تاريخ الطباعة' if rtl else 'Generated on'
        elements.append(Paragraph(f"{gen_label}: {date_str}", subtitle_style))
    
    # Separator line
    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E5E7EB'), spaceAfter=12))
    
    # Chart image (if provided)
    if chart_image:
        try:
            chart_buffer = io.BytesIO(chart_image)
            chart_width = page_size[0] - 120
            chart_height = chart_width * 0.5
            img = Image(chart_buffer, width=chart_width, height=chart_height)
            elements.append(img)
            elements.append(Spacer(1, 16))
        except Exception as e:
            logger.warning(f"Failed to embed chart image: {e}")
    
    # Prepare Table Data
    table_data = []
    
    if columns:
        header_row = [_process_cell(c, rtl and arabic_available) for c in columns]
        if rtl:
            header_row = list(reversed(header_row))
        table_data.append(header_row)
    elif data and isinstance(data[0], list):
        header_row = [_process_cell(c, rtl and arabic_available) for c in data[0]]
        if rtl:
            header_row = list(reversed(header_row))
        table_data.append(header_row)
        data = data[1:]
        if columns is None:
            columns = data[0] if not table_data else None
    elif data and isinstance(data[0], dict):
        hdr = list(data[0].keys())
        header_row = [_process_cell(c, rtl and arabic_available) for c in hdr]
        if rtl:
            header_row = list(reversed(header_row))
        table_data.append(header_row)
        columns = hdr
    
    # Rows
    for row in data:
        if isinstance(row, dict):
            col_keys = columns or list(row.keys())
            cells = [_process_cell(row.get(c, ''), rtl and arabic_available) for c in col_keys]
        elif isinstance(row, list):
            cells = [_process_cell(c, rtl and arabic_available) for c in row]
        else:
            continue
        if rtl:
            cells = list(reversed(cells))
        table_data.append(cells)
    
    if not table_data:
        elements.append(Paragraph("لا توجد بيانات / No data" if rtl else "No data", normal_style))
    else:
        col_count = len(table_data[0])
        available_width = page_size[0] - 80
        col_width = available_width / max(col_count, 1)
        col_widths = [col_width] * col_count

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        header_bg = colors.HexColor('#1E3A5F')
        alt_row_bg = colors.HexColor('#F8FAFC')
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), header_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_row_bg]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT' if rtl else 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]
        
        # Detect numeric columns for center/right alignment
        for col_idx in range(col_count):
            numeric_count = 0
            for row_idx in range(1, min(len(table_data), 6)):
                try:
                    val = str(table_data[row_idx][col_idx]).replace(',', '').replace('%', '').strip()
                    float(val)
                    numeric_count += 1
                except (ValueError, IndexError):
                    pass
            if numeric_count >= min(len(table_data) - 1, 3):
                table_style.append(('ALIGN', (col_idx, 0), (col_idx, -1), 'CENTER'))
        
        t.setStyle(TableStyle(table_style))
        elements.append(t)
    
    doc.build(elements, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    buffer.seek(0)
    return buffer


# ─── Chart Generation (RPT-102) ─────────────────────────────────────────────

def generate_chart_image(chart_type: str, labels: list, datasets: list,
                         title: str = "", rtl: bool = False,
                         width: int = 800, height: int = 400) -> bytes:
    """
    Generate a chart image (PNG bytes) using matplotlib.
    
    chart_type: "bar", "line", "pie", "horizontal_bar"
    labels: X-axis labels (list of strings)
    datasets: List of dicts: [{"label": "Revenue", "data": [100, 200, 300], "color": "#2563eb"}]
    title: Chart title
    rtl: If True, set Arabic-friendly rendering
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
        from matplotlib import rcParams
        
        # Arabic font support
        if rtl:
            try:
                from matplotlib.font_manager import FontProperties
                arabic_font = FontProperties(fname=ARABIC_FONT_PATH) if os.path.exists(ARABIC_FONT_PATH) else None
            except Exception:
                arabic_font = None
        else:
            arabic_font = None
        
        fig, ax = plt.subplots(figsize=(width / 100, height / 100), dpi=100)
        fig.patch.set_facecolor('#FFFFFF')
        ax.set_facecolor('#FAFAFA')
        
        colors_palette = ['#2563EB', '#16A34A', '#DC2626', '#7C3AED', '#D97706',
                          '#06B6D4', '#EC4899', '#8B5CF6', '#14B8A6', '#F43F5E']
        
        if chart_type == "bar":
            import numpy as np
            x = np.arange(len(labels))
            bar_width = 0.8 / max(len(datasets), 1)
            for i, ds in enumerate(datasets):
                color = ds.get("color", colors_palette[i % len(colors_palette)])
                offset = (i - len(datasets) / 2 + 0.5) * bar_width
                bars = ax.bar(x + offset, ds["data"], bar_width, label=ds.get("label", ""),
                             color=color, edgecolor='white', linewidth=0.5)
                # Value labels on bars
                for bar_item in bars:
                    height_val = bar_item.get_height()
                    if height_val > 0:
                        ax.text(bar_item.get_x() + bar_item.get_width() / 2., height_val,
                                f'{height_val:,.0f}', ha='center', va='bottom', fontsize=7)
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45 if len(labels) > 6 else 0, ha='right', fontsize=8)
            
        elif chart_type == "horizontal_bar":
            import numpy as np
            y = np.arange(len(labels))
            for i, ds in enumerate(datasets):
                color = ds.get("color", colors_palette[i % len(colors_palette)])
                ax.barh(y, ds["data"], label=ds.get("label", ""), color=color, edgecolor='white')
            ax.set_yticks(y)
            ax.set_yticklabels(labels, fontsize=8)
            
        elif chart_type == "line":
            for i, ds in enumerate(datasets):
                color = ds.get("color", colors_palette[i % len(colors_palette)])
                ax.plot(labels, ds["data"], marker='o', linewidth=2, markersize=4,
                       label=ds.get("label", ""), color=color)
                # Fill area
                ax.fill_between(range(len(labels)), ds["data"], alpha=0.08, color=color)
            ax.set_xticklabels(labels, rotation=45 if len(labels) > 6 else 0, ha='right', fontsize=8)
            
        elif chart_type == "pie":
            ds = datasets[0] if datasets else {"data": [], "label": ""}
            pie_colors = [ds.get("color", colors_palette[i % len(colors_palette)]) 
                         for i in range(len(ds["data"]))]
            wedges, texts, autotexts = ax.pie(
                ds["data"], labels=labels, autopct='%1.1f%%',
                colors=pie_colors, startangle=90, pctdistance=0.85
            )
            for text in autotexts:
                text.set_fontsize(8)
            centre_circle = plt.Circle((0, 0), 0.55, fc='white')
            ax.add_artist(centre_circle)
        
        if title:
            if arabic_font:
                ax.set_title(title, fontproperties=arabic_font, fontsize=13, pad=15, color='#1F2937')
            else:
                ax.set_title(title, fontsize=13, pad=15, fontweight='bold', color='#1F2937')
        
        if chart_type != "pie" and len(datasets) > 1:
            ax.legend(fontsize=8, loc='upper right', framealpha=0.9)
        
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#E5E7EB')
        ax.spines['bottom'].set_color('#E5E7EB')
        ax.tick_params(colors='#6B7280', labelsize=8)
        ax.grid(axis='y', alpha=0.3, color='#E5E7EB')
        
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight',
                    facecolor='white', edgecolor='none')
        plt.close(fig)
        img_buffer.seek(0)
        return img_buffer.read()
    except Exception as e:
        logger.error(f"Chart generation failed: {e}")
        return None


def generate_excel_with_chart(data: list, columns: list, sheet_name: str = "Report",
                              chart_type: str = None, chart_config: dict = None) -> io.BytesIO:
    """
    Generate Excel file with optional embedded chart.
    
    chart_type: "bar", "line", "pie" (openpyxl chart types)
    chart_config: {"title": "...", "x_col": 0, "y_cols": [1,2], "chart_row": len(data)+3}
    """
    df = pd.DataFrame(data)
    
    if columns:
        valid_cols = [c for c in columns if c in df.columns]
        if valid_cols:
            df = df[valid_cols]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        worksheet = writer.sheets[sheet_name]
        
        # Auto-adjust column width
        for idx, col in enumerate(df.columns):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(col))
            )) + 2
            col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + idx % 26)
            worksheet.column_dimensions[col_letter].width = min(max_len, 50)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Style data rows with alternating colors
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        for row_idx in range(2, len(df) + 2):
            for cell in worksheet[row_idx]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        
        # Add embedded chart if requested
        if chart_type and chart_config and len(df) > 0:
            try:
                from openpyxl.chart import BarChart, LineChart, PieChart, Reference
                
                num_rows = len(df) + 1
                ChartClass = {'bar': BarChart, 'line': LineChart, 'pie': PieChart}.get(chart_type, BarChart)
                chart = ChartClass()
                chart.title = chart_config.get("title", "")
                chart.width = 20
                chart.height = 12
                chart.style = 10
                
                x_col = chart_config.get("x_col", 0) + 1
                y_cols = chart_config.get("y_cols", [1])
                
                categories = Reference(worksheet, min_col=x_col, min_row=2, max_row=num_rows)
                
                for y_col in y_cols:
                    values = Reference(worksheet, min_col=y_col + 1, min_row=1, max_row=num_rows)
                    chart.add_data(values, titles_from_data=True)
                
                chart.set_categories(categories)
                
                chart_row = chart_config.get("chart_row", num_rows + 3)
                worksheet.add_chart(chart, f"A{chart_row}")
            except Exception as e:
                logger.warning(f"Failed to add chart to Excel: {e}")
    
    output.seek(0)
    return output

def create_export_response(file_buffer: io.BytesIO, filename: str, content_type: str):
    return StreamingResponse(
        file_buffer,
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
